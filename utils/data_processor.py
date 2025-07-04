#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
数据处理工具
"""

import pandas as pd
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


class DataProcessor:
    """数据处理工具类"""
    
    def __init__(self):
        """初始化数据处理器"""
        pass
    
    def deduplicate_papers(self, df, verbose=True):
        """
        对论文进行去重处理
        
        Args:
            df: 包含论文信息的DataFrame
            verbose: 是否打印详细信息
        
        Returns:
            DataFrame: 去重后的论文数据
        """
        if verbose:
            print("开始论文去重处理...")
        
        original_count = len(df)
        
        # 第一阶段：根据Title、Year、Author、Journal、Booktitle列去掉完全一致的论文
        if verbose:
            print(f"第一阶段：根据完全匹配去重（Title, Year, Author, Journal, Booktitle）")
        
        df_stage1 = df.drop_duplicates(subset=['Title', 'Year', 'Author', 'Journal', 'Booktitle'])
        stage1_removed = original_count - len(df_stage1)
        
        if verbose:
            print(f"  - 原始论文数量: {original_count}")
            print(f"  - 去重后数量: {len(df_stage1)}")
            print(f"  - 去除重复: {stage1_removed} 篇")
        
        # 第二阶段：在Title重复的论文中，保留非CORR的版本
        if verbose:
            print(f"第二阶段：优先保留非CORR版本的论文")
        
        # 找出所有Title重复的论文
        duplicate_titles = df_stage1[df_stage1.duplicated(subset=['Title'], keep=False)]
        
        if verbose and len(duplicate_titles) > 0:
            print(f"  - 发现 {len(duplicate_titles)} 篇同标题论文")
        
        # 对于重复的论文，保留非CORR的版本
        deduplicated_papers = []
        
        for title in df_stage1['Title'].unique():
            same_title_papers = df_stage1[df_stage1['Title'] == title]
            
            if len(same_title_papers) > 1:
                # 如果有多篇同标题论文，优先保留非CORR版本
                non_corr_papers = same_title_papers[same_title_papers['Journal'] != 'CoRR']
                
                if len(non_corr_papers) > 0:
                    # 有非CORR版本，保留第一个非CORR版本
                    deduplicated_papers.append(non_corr_papers.iloc[0])
                    if verbose and len(same_title_papers) > 1:
                        corr_count = len(same_title_papers[same_title_papers['Journal'] == 'CoRR'])
                        if corr_count > 0:
                            print(f"  - 保留非CORR版本: {title[:50]}... (去除{corr_count}个CORR版本)")
                else:
                    # 都是CORR版本，保留第一个
                    deduplicated_papers.append(same_title_papers.iloc[0])
                    if verbose:
                        print(f"  - 保留CORR版本: {title[:50]}... (去除{len(same_title_papers)-1}个重复CORR版本)")
            else:
                # 只有一篇论文，直接保留
                deduplicated_papers.append(same_title_papers.iloc[0])
        
        df_final = pd.DataFrame(deduplicated_papers)
        stage2_removed = len(df_stage1) - len(df_final)
        
        if verbose:
            print(f"  - 第二阶段前数量: {len(df_stage1)}")
            print(f"  - 第二阶段后数量: {len(df_final)}")
            print(f"  - 去除重复: {stage2_removed} 篇")
            print(f"论文去重完成！总共去除 {original_count - len(df_final)} 篇重复论文")
        
        return df_final
    
    def create_journal_ranking_excel(self, results, journal_counts, summary, output_file):
        """
        创建期刊分区结果Excel文件
        
        Args:
            results: 匹配结果字典
            journal_counts: 期刊论文数量字典
            summary: 统计摘要
            output_file: 输出文件路径
        """
        # 创建工作簿
        wb = Workbook()
        
        # 删除默认工作表
        wb.remove(wb.active)
        
        # 创建详细结果表
        self._create_detailed_results_sheet(wb, results, journal_counts)
        
        # 创建统计摘要表
        self._create_summary_sheet(wb, summary)
        
        # 创建分区分布表
        self._create_distribution_sheet(wb, results, journal_counts)
        
        # 保存文件
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        wb.save(output_file)
    
    def _create_detailed_results_sheet(self, wb, results, journal_counts):
        """创建详细结果工作表"""
        ws = wb.create_sheet("详细结果", 0)
        
        # 设置标题
        headers = [
            '期刊缩写', '论文数量', '期刊标题', 'ISSN列表',
            'CCF匹配', 'CCF期刊名', 'CCF分区', 'CCF类型',
            '中科院匹配', '中科院期刊名', '中科院分区', '中科院大类', '匹配方式'
        ]
        
        # 写入标题
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='B0C4DE', end_color='B0C4DE', fill_type='solid')
        
        # 写入数据
        row = 2
        for journal_abbr, result in results.items():
            paper_count = journal_counts.get(journal_abbr, 0)
            
            ws.cell(row=row, column=1, value=journal_abbr)
            ws.cell(row=row, column=2, value=paper_count)
            ws.cell(row=row, column=3, value=result.get('title', ''))
            ws.cell(row=row, column=4, value='; '.join(result.get('issn_list', [])))
            
            # CCF信息
            ccf = result.get('ccf', {})
            ws.cell(row=row, column=5, value='是' if ccf.get('matched') else '否')
            ws.cell(row=row, column=6, value=ccf.get('name', ''))
            ws.cell(row=row, column=7, value=ccf.get('rank', ''))
            ws.cell(row=row, column=8, value=ccf.get('type', ''))
            
            # 中科院信息
            cas = result.get('cas', {})
            ws.cell(row=row, column=9, value='是' if cas.get('matched') else '否')
            ws.cell(row=row, column=10, value=cas.get('name', ''))
            ws.cell(row=row, column=11, value=cas.get('zone', ''))
            ws.cell(row=row, column=12, value=cas.get('category', ''))
            ws.cell(row=row, column=13, value=cas.get('match_type', ''))
            
            # 设置行高
            ws.row_dimensions[row].height = 20
            
            row += 1
        
        # 调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15
    
    def _create_summary_sheet(self, wb, summary):
        """创建统计摘要工作表"""
        ws = wb.create_sheet("统计摘要", 1)
        
        # 标题
        ws.cell(row=1, column=1, value="期刊分区统计摘要")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        # 基本统计
        row = 3
        ws.cell(row=row, column=1, value="总期刊数量")
        # 兼容两种键名：total_journals（旧版本）和total_publications（新版本）
        total_count = summary.get('total_journals', summary.get('total_publications', 0))
        ws.cell(row=row, column=2, value=total_count)
        row += 1
        
        ws.cell(row=row, column=1, value="CCF匹配数量")
        ws.cell(row=row, column=2, value=summary['ccf_matches'])
        row += 1
        
        ws.cell(row=row, column=1, value="中科院匹配数量")
        ws.cell(row=row, column=2, value=summary['cas_matches'])
        row += 2
        
        # CCF分区统计
        ws.cell(row=row, column=1, value="CCF分区统计")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        for rank, count in summary['ccf_stats'].items():
            if count > 0:
                ws.cell(row=row, column=1, value=f"{rank}类期刊")
                ws.cell(row=row, column=2, value=count)
                ws.cell(row=row, column=3, value=f"{summary['ccf_papers'][rank]}篇论文")
                row += 1
        
        row += 1
        
        # 中科院分区统计
        ws.cell(row=row, column=1, value="中科院分区统计")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        for zone, count in summary['cas_stats'].items():
            if count > 0:
                ws.cell(row=row, column=1, value=f"{zone}期刊")
                ws.cell(row=row, column=2, value=count)
                ws.cell(row=row, column=3, value=f"{summary['cas_papers'][zone]}篇论文")
                row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def _create_distribution_sheet(self, wb, results, journal_counts):
        """创建分区分布工作表"""
        ws = wb.create_sheet("分区分布", 2)
        
        # CCF分区分布
        ws.cell(row=1, column=1, value="CCF分区分布")
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        
        row = 3
        for rank in ['A', 'B', 'C']:
            ws.cell(row=row, column=1, value=f"CCF {rank}类")
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            
            for journal_abbr, result in results.items():
                ccf = result.get('ccf', {})
                if ccf.get('matched') and ccf.get('rank') == rank:
                    paper_count = journal_counts.get(journal_abbr, 0)
                    ws.cell(row=row, column=2, value=journal_abbr)
                    ws.cell(row=row, column=3, value=paper_count)
                    ws.cell(row=row, column=4, value=ccf.get('name', ''))
                    row += 1
            
            row += 1
        
        # 中科院分区分布
        row += 2
        ws.cell(row=row, column=1, value="中科院分区分布")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 2
        
        for zone in ['1区', '2区', '3区', '4区']:
            ws.cell(row=row, column=1, value=f"中科院 {zone}")
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            
            for journal_abbr, result in results.items():
                cas = result.get('cas', {})
                if cas.get('matched') and cas.get('zone') == zone:
                    paper_count = journal_counts.get(journal_abbr, 0)
                    ws.cell(row=row, column=2, value=journal_abbr)
                    ws.cell(row=row, column=3, value=paper_count)
                    ws.cell(row=row, column=4, value=cas.get('name', ''))
                    row += 1
            
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 30
    
    def save_summary_json(self, summary, output_file):
        """保存统计摘要为JSON文件"""
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(summary, f, ensure_ascii=False, indent=2)
    
    def print_summary(self, summary):
        """打印统计摘要"""
        print("\n=== 期刊分区统计摘要 ===")
        # 兼容两种键名：total_journals（旧版本）和total_publications（新版本）
        total_count = summary.get('total_journals', summary.get('total_publications', 0))
        print(f"总期刊数量: {total_count}")
        if total_count > 0:
            print(f"CCF匹配成功: {summary['ccf_matches']} ({summary['ccf_matches']/total_count*100:.1f}%)")
            print(f"中科院匹配成功: {summary['cas_matches']} ({summary['cas_matches']/total_count*100:.1f}%)")
        else:
            print(f"CCF匹配成功: {summary['ccf_matches']}")
            print(f"中科院匹配成功: {summary['cas_matches']}")
        
        print("\n=== CCF分区分布 ===")
        for rank, count in summary['ccf_stats'].items():
            if count > 0:
                print(f"{rank}类: {count}个期刊, {summary['ccf_papers'][rank]}篇论文")
        
        print("\n=== 中科院分区分布 ===")
        for zone, count in summary['cas_stats'].items():
            if count > 0:
                print(f"{zone}: {count}个期刊, {summary['cas_papers'][zone]}篇论文")
        
        print("\n=== 重点分区论文数量 ===")
        print(f"CCF A类论文: {summary['ccf_papers']['A']}篇")
        print(f"CCF B类论文: {summary['ccf_papers']['B']}篇")
        print(f"中科院1区论文: {summary['cas_papers']['1区']}篇")
        print(f"中科院2区论文: {summary['cas_papers']['2区']}篇") 

    def add_ranking_info_to_papers(self, df, ranking_results):
        """
        为论文DataFrame添加CCF和中科院分区信息
        
        Args:
            df: 论文DataFrame
            ranking_results: 分区匹配结果字典
            
        Returns:
            DataFrame: 添加了分区信息的DataFrame
        """
        print("为论文添加分区信息...")
        df_with_rankings = df.copy()
        
        # 添加分区字段
        df_with_rankings['CCF_Rank'] = ''
        df_with_rankings['CCF_Name'] = ''
        df_with_rankings['CAS_Zone'] = ''
        df_with_rankings['CAS_Top'] = ''
        df_with_rankings['CAS_Name'] = ''
        df_with_rankings['Publication_Type'] = ''
        
        # 统计信息
        ccf_matched_count = 0
        cas_matched_count = 0
        processed_count = 0
        
        # 为每篇论文匹配分区信息
        for idx, row in df_with_rankings.iterrows():
            # 确定论文类型和出版物名称
            if row['Type'] == 'article' and row['Journal']:
                pub_type = 'journal'
                pub_name = row['Journal']
            elif row['Type'] == 'inproceedings' and row['Booktitle']:
                pub_type = 'conference'
                pub_name = row['Booktitle']
            else:
                continue
            
            df_with_rankings.at[idx, 'Publication_Type'] = pub_type
            processed_count += 1
            
            # 从DBLP ID中提取出版物缩写
            dblp_id = row['ID']
            if dblp_id and 'DBLP:' in dblp_id:
                try:
                    if pub_type == 'journal' and 'journals/' in dblp_id:
                        # 提取期刊缩写：DBLP:journals/abbr/... -> abbr
                        raw_abbr = dblp_id.split('journals/')[1].split('/')[0]
                        # 构造带前缀的键
                        abbr_key = f"journal_{raw_abbr}"
                    elif pub_type == 'conference' and 'conf/' in dblp_id:
                        # 提取会议缩写：DBLP:conf/abbr/... -> abbr
                        raw_abbr = dblp_id.split('conf/')[1].split('/')[0]
                        # 构造带前缀的键
                        abbr_key = f"conference_{raw_abbr}"
                    else:
                        continue
                        
                    # 在排名结果中查找匹配
                    if abbr_key in ranking_results:
                        result = ranking_results[abbr_key]
                        
                        # 添加CCF分级信息
                        ccf_result = result.get('ccf', {})
                        if ccf_result.get('matched', False):
                            df_with_rankings.at[idx, 'CCF_Rank'] = ccf_result.get('rank', '')
                            df_with_rankings.at[idx, 'CCF_Name'] = ccf_result.get('name', '')
                            df_with_rankings.at[idx, 'CCF_Type'] = ccf_result.get('type', '')
                            ccf_matched_count += 1
                        
                        # 添加中科院分区信息（仅期刊）
                        if pub_type == 'journal':
                            cas_result = result.get('cas', {})
                            if cas_result.get('matched', False):
                                df_with_rankings.at[idx, 'CAS_Zone'] = cas_result.get('zone', '')
                                df_with_rankings.at[idx, 'CAS_Name'] = cas_result.get('name', '')
                                df_with_rankings.at[idx, 'CAS_Top'] = cas_result.get('top', '')
                                df_with_rankings.at[idx, 'CAS_Category'] = cas_result.get('category', '')
                                cas_matched_count += 1
                    elif raw_abbr in ranking_results:
                        # 向后兼容：如果找不到带前缀的键，尝试原始缩写
                        result = ranking_results[raw_abbr]
                        
                        # 添加CCF分级信息
                        ccf_result = result.get('ccf', {})
                        if ccf_result.get('matched', False):
                            df_with_rankings.at[idx, 'CCF_Rank'] = ccf_result.get('rank', '')
                            df_with_rankings.at[idx, 'CCF_Name'] = ccf_result.get('name', '')
                            df_with_rankings.at[idx, 'CCF_Type'] = ccf_result.get('type', '')
                            ccf_matched_count += 1
                        
                        # 添加中科院分区信息（仅期刊）
                        if pub_type == 'journal':
                            cas_result = result.get('cas', {})
                            if cas_result.get('matched', False):
                                df_with_rankings.at[idx, 'CAS_Zone'] = cas_result.get('zone', '')
                                df_with_rankings.at[idx, 'CAS_Name'] = cas_result.get('name', '')
                                df_with_rankings.at[idx, 'CAS_Top'] = cas_result.get('top', '')
                                df_with_rankings.at[idx, 'CAS_Category'] = cas_result.get('category', '')
                                cas_matched_count += 1
                        
                except Exception as e:
                    print(f"处理论文 {idx} 时出错: {e}")
                    continue
        
        print(f"分区信息添加完成:")
        print(f"  处理论文数: {processed_count}")
        print(f"  CCF匹配数: {ccf_matched_count}")
        print(f"  中科院匹配数: {cas_matched_count}")
        
        return df_with_rankings 

 

    def analyze_paper_rankings(self, df_with_rankings, verbose=True):
        """
        分析论文分区统计
        
        Args:
            df_with_rankings: 带有分区信息的论文DataFrame
            verbose: 是否打印详细信息
            
        Returns:
            dict: 论文分区统计结果
        """
        if verbose:
            print("=== 论文分区统计分析 ===")
        
        # 基本统计
        total_papers = len(df_with_rankings)
        
        # CCF分区统计
        ccf_a_papers = df_with_rankings[df_with_rankings['CCF_Rank'] == 'A类']
        ccf_b_papers = df_with_rankings[df_with_rankings['CCF_Rank'] == 'B类']
        ccf_c_papers = df_with_rankings[df_with_rankings['CCF_Rank'] == 'C类']
        
        # 中科院分区统计
        cas_1_papers = df_with_rankings[df_with_rankings['CAS_Zone'] == '1区']
        cas_2_papers = df_with_rankings[df_with_rankings['CAS_Zone'] == '2区']
        cas_3_papers = df_with_rankings[df_with_rankings['CAS_Zone'] == '3区']
        cas_4_papers = df_with_rankings[df_with_rankings['CAS_Zone'] == '4区']
        
        # 中科院TOP期刊论文统计
        cas_top_papers = df_with_rankings[df_with_rankings['CAS_Top'] == '是']
        
        # 组合统计（去重）
        # 1. CCF-A + 中科院一区 (并集)
        ccf_a_or_cas_1 = df_with_rankings[
            (df_with_rankings['CCF_Rank'] == 'A类') | 
            (df_with_rankings['CAS_Zone'] == '1区')
        ]
        
        # 2. CCF-A + CCF-B + 中科院一区和二区 (并集)
        ccf_ab_or_cas_12 = df_with_rankings[
            (df_with_rankings['CCF_Rank'].isin(['A类', 'B类'])) |
            (df_with_rankings['CAS_Zone'].isin(['1区', '2区']))
        ]
        
        # 构建统计结果
        stats = {
            'total_papers': total_papers,
            
            # CCF分区统计
            'ccf_a_count': len(ccf_a_papers),
            'ccf_b_count': len(ccf_b_papers),
            'ccf_c_count': len(ccf_c_papers),
            
            # 中科院分区统计
            'cas_1_count': len(cas_1_papers),
            'cas_2_count': len(cas_2_papers),
            'cas_3_count': len(cas_3_papers),
            'cas_4_count': len(cas_4_papers),
            'cas_top_count': len(cas_top_papers),
            
            # 组合统计
            'ccf_a_or_cas_1_count': len(ccf_a_or_cas_1),
            'ccf_ab_or_cas_12_count': len(ccf_ab_or_cas_12),
            
            # 计算比例
            'ccf_a_ratio': (len(ccf_a_papers) / total_papers * 100) if total_papers > 0 else 0,
            'ccf_b_ratio': (len(ccf_b_papers) / total_papers * 100) if total_papers > 0 else 0,
            'ccf_c_ratio': (len(ccf_c_papers) / total_papers * 100) if total_papers > 0 else 0,
            'cas_1_ratio': (len(cas_1_papers) / total_papers * 100) if total_papers > 0 else 0,
            'cas_2_ratio': (len(cas_2_papers) / total_papers * 100) if total_papers > 0 else 0,
            'cas_3_ratio': (len(cas_3_papers) / total_papers * 100) if total_papers > 0 else 0,
            'cas_4_ratio': (len(cas_4_papers) / total_papers * 100) if total_papers > 0 else 0,
            'cas_top_ratio': (len(cas_top_papers) / total_papers * 100) if total_papers > 0 else 0,
            'ccf_a_or_cas_1_ratio': (len(ccf_a_or_cas_1) / total_papers * 100) if total_papers > 0 else 0,
            'ccf_ab_or_cas_12_ratio': (len(ccf_ab_or_cas_12) / total_papers * 100) if total_papers > 0 else 0,
            
            # 保存数据框以供后续使用
            'ccf_a_papers': ccf_a_papers,
            'ccf_b_papers': ccf_b_papers,
            'ccf_c_papers': ccf_c_papers,
            'cas_1_papers': cas_1_papers,
            'cas_2_papers': cas_2_papers,
            'cas_3_papers': cas_3_papers,
            'cas_4_papers': cas_4_papers,
            'cas_top_papers': cas_top_papers,
            'ccf_a_or_cas_1_papers': ccf_a_or_cas_1,
            'ccf_ab_or_cas_12_papers': ccf_ab_or_cas_12,
        }
        
        if verbose:
            self.print_ranking_report(stats)
            
        return stats
    
    def print_ranking_report(self, stats):
        """
        打印论文分区统计报告
        
        Args:
            stats: 统计结果字典
        """
        print(f"\n📊 总论文数: {stats['total_papers']}")
        print(f"\n=== CCF分区统计 ===")
        print(f"🅰️  CCF-A类论文: {stats['ccf_a_count']}篇 ({stats['ccf_a_ratio']:.1f}%)")
        print(f"🅱️  CCF-B类论文: {stats['ccf_b_count']}篇 ({stats['ccf_b_ratio']:.1f}%)")
        print(f"🅲  CCF-C类论文: {stats['ccf_c_count']}篇 ({stats['ccf_c_ratio']:.1f}%)")
        
        print(f"\n=== 中科院分区统计 ===")
        print(f"🥇 中科院1区论文: {stats['cas_1_count']}篇 ({stats['cas_1_ratio']:.1f}%)")
        print(f"⭐ 中科院TOP论文: {stats['cas_top_count']}篇 ({stats['cas_top_ratio']:.1f}%)")

        print(f"🥈 中科院2区论文: {stats['cas_2_count']}篇 ({stats['cas_2_ratio']:.1f}%)")
        print(f"🥉 中科院3区论文: {stats['cas_3_count']}篇 ({stats['cas_3_ratio']:.1f}%)")
        print(f"🏅 中科院4区论文: {stats['cas_4_count']}篇 ({stats['cas_4_ratio']:.1f}%)")
   
        
        print(f"\n=== 组合统计 (去重) ===")
        print(f"🏆 CCF-A类 + 中科院1区 (并集): {stats['ccf_a_or_cas_1_count']}篇 ({stats['ccf_a_or_cas_1_ratio']:.1f}%)")
        print(f"🌟 CCF-A/B类 + 中科院1/2区 (并集): {stats['ccf_ab_or_cas_12_count']}篇 ({stats['ccf_ab_or_cas_12_ratio']:.1f}%)")
        

        
        # 显示CCF-A类论文样例
        if stats['ccf_a_count'] > 0:
            print(f"\n=== CCF-A类论文样例 ===")
            sample_count = min(3, stats['ccf_a_count'])
            for i, (idx, row) in enumerate(stats['ccf_a_papers'].head(sample_count).iterrows()):
                print(f"{i+1}. {row['Title'][:60]}...")
                pub_name = row['Journal'] if row['Journal'] else row['Booktitle']
                print(f"   发表于: {pub_name}")
        
        if stats['cas_1_count'] > 0:
            print(f"\n=== 中科院1区论文样例 ===")
            sample_count = min(3, stats['cas_1_count'])
            for i, (idx, row) in enumerate(stats['cas_1_papers'].head(sample_count).iterrows()):
                print(f"{i+1}. {row['Title'][:60]}...")
                print(f"   发表于: {row['Journal']}") 

    def generate_scholar_summary(self, df_with_rankings, print_summary=True):
        """
        生成论文分区统计摘要
        
        Args:
            df_with_rankings: 带有分区信息的论文DataFrame
            print_summary: 是否打印摘要，默认True
            
        Returns:
            dict: 包含论文分区统计信息的字典
        """
        # 获取详细统计
        ranking_stats = self.analyze_paper_rankings(df_with_rankings, verbose=False)
        
        # 生成简洁摘要
        summary = {
            "总论文数": ranking_stats['total_papers'],
            "CCF-A类+中科院1区": ranking_stats['ccf_a_or_cas_1_count'],  # CCF-A类+中科院1区
            "CCF-A/B类+中科院1/2区": ranking_stats['ccf_ab_or_cas_12_count'],  # CCF-A/B类+中科院1/2区
            "CCF分区": {
                "A类": ranking_stats['ccf_a_count'],
                "B类": ranking_stats['ccf_b_count'],
                "C类": ranking_stats['ccf_c_count']
            },
            "中科院分区": {
                "1区": ranking_stats['cas_1_count'],
                "2区": ranking_stats['cas_2_count'],
                "3区": ranking_stats['cas_3_count'],
                "4区": ranking_stats['cas_4_count']
            },
        }
        
        if print_summary:
            print("=== 论文分区统计摘要 ===")
            print(f"📊 总论文数: {summary['总论文数']}")
            print(f"🏆 CCF-A类+中科院1区: {summary['CCF-A类+中科院1区']}篇")
            print(f"🌟 CCF-A/B类+中科院1/2区: {summary['CCF-A/B类+中科院1/2区']}篇")
            print(f"🎯 CCF分区: A类 {summary['CCF分区']['A类']}篇 | B类 {summary['CCF分区']['B类']}篇 | C类 {summary['CCF分区']['C类']}篇")
            print(f"🔬 中科院分区: 1区 {summary['中科院分区']['1区']}篇 | 2区 {summary['中科院分区']['2区']}篇 | 3区 {summary['中科院分区']['3区']}篇 | 4区 {summary['中科院分区']['4区']}篇")
            
        return summary
    
 